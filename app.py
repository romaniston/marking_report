from flask import Flask, render_template, request
from openpyxl import load_workbook
from datetime import datetime, date
from collections import defaultdict
from openpyxl.utils.datetime import from_excel

app = Flask(__name__)

@app.route('/')
def index():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    masters_data = defaultdict(int)
    result = None  # <- Пусто по умолчанию, если даты не заданы

    wb = load_workbook('tables/test_table.xlsx', data_only=True)
    ws = wb.active

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_col == 16:
                    row = merged_range.min_row

                    cell_date_raw = ws.cell(row=row, column=3).value

                    if isinstance(cell_date_raw, datetime):
                        cell_date = cell_date_raw.date()
                    elif isinstance(cell_date_raw, date):
                        cell_date = cell_date_raw
                    elif isinstance(cell_date_raw, (int, float)):
                        try:
                            cell_date = from_excel(cell_date_raw).date()
                        except:
                            continue
                    else:
                        continue

                    print(f'cell_date_raw: {cell_date_raw}, cell_date: {cell_date}')

                    if not (start_date <= cell_date <= end_date):
                        continue

                    master_raw = ws.cell(row=row, column=5).value
                    try:
                        master = str(master_raw.replace(" ", "").replace(".", "").lower())
                    except Exception:
                        pass

                    output = ws.cell(row=row, column=16).value

                    print(f'master: {master}, output: {output}')                    

                    if master and output:
                        masters_data[master] += output

            result = [{'name': name, 'output': output} for name, output in masters_data.items()]

        except Exception as e:
            print("Ошибка обработки:", e)

    return render_template('index.html', masters=result)

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=4995, debug=True)
